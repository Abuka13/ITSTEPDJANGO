from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import load_workbook





def home(request):
    file_path = 'data.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active
    matrix = []
    for row in sheet.iter_rows(values_only=True):
        matrix.append(row)
    return render(request, "home.html", context={"data":matrix})